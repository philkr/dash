import datetime
from typing import Dict, List
import logging
from pathlib import Path
from contextlib import contextmanager
import re


class Entry:
    name: str
    timestamp: datetime.datetime
    value: float

    def __init__(self, name: str, value: float, timestamp: datetime.datetime):
        self.name = name
        self.value = value
        self.timestamp = timestamp or datetime.datetime.now()


class TableSettings(object):
    """ Hide a table in the visualization """
    hide: bool = False
    """ Axis label """
    axis: str = ''
    """ Plot name """
    name: str = ''
    """ Priority to sort plots with, higher priority plots are shown earlier """
    priority: int = 10
    """ Keep top n results """
    keep_top: int = 5
    """ Keep last n results """
    keep_last: int = 5
    """ Is a lower value better? """
    lower_better: bool = False


class Table:
    baselines: List[Entry]
    entries: List[Entry]
    settings: TableSettings

    def __init__(self):
        self.baselines = []
        self.entries = []
        self.settings = TableSettings()

    def add(self, e: Entry):
        self.entries.append(e)

    def add_baseline(self, e: Entry):
        self.baselines.append(e)

    def setting(self, **kwargs):
        for a, v in kwargs.items():
            if not hasattr(self.settings, a):
                logging.warning('Setting "%s" not found!', a)
            else:
                T = type(getattr(self.settings, a))
                setattr(self.settings, a, T(v))


class Interface(object):
    name: str

    def __init__(self, name: str):
        self.name = name

    def poll(self):
        pass


class MultiInterface(object):
    interfaces: List[Interface]

    def __init__(self):
        self.interfaces = []

    def poll(self):
        pass

    def __len__(self):
        return len(self.interfaces)

    def __getitem__(self, item):
        return self.interfaces[item]


class TabularInterface(Interface):
    tables: List[Table]

    def __init__(self, name: str):
        super().__init__(name)
        self.tables = []


class BaseSheetInterface(TabularInterface):
    SETTING_RE = re.compile(r's:(\w+)(=?)(.*)')
    BASELINE_RE = re.compile(r'b:(.*)')
    IGNORE_RE = re.compile(r'i:(.*)')

    @contextmanager
    def file(self):
        raise NotImplementedError

    @staticmethod
    def _prase_worksheet(ws):
        tb = Table()
        tb.setting(name=ws.title)
        for r in ws.rows:
            name, ts, value, baseline, ignore = '', None, None, False, False
            for c in r:
                if c.value is not None:
                    if c.is_date:
                        ts = c.value
                    elif isinstance(c.value, (float, int)):
                        value = float(c.value)
                        # Read only one value
                        break
                    else:
                        m_s = BaseSheetInterface.SETTING_RE.match(c.value)
                        m_b = BaseSheetInterface.BASELINE_RE.match(c.value)
                        m_i = BaseSheetInterface.IGNORE_RE.match(c.value)
                        if m_s:
                            if m_s.group(2):
                                tb.setting(**{m_s.group(1): m_s.group(3)})
                            else:
                                tb.setting(**{m_s.group(1): True})
                        elif m_b:
                            baseline = True
                            name += m_b.group(1)
                        elif m_i:
                            ignore = True
                        else:
                            name += c.value
            if not ignore and value is not None and name:
                if baseline:
                    tb.add_baseline(Entry(name, value, ts))
                else:
                    tb.add(Entry(name, value, ts))
        return tb

    def poll(self):
        import openpyxl
        self.tables = []
        with self.file() as f:
            wb = openpyxl.load_workbook(f)
            for ws in wb.worksheets:
                self.tables.append(self._prase_worksheet(ws))


class LocalSheetInterface(BaseSheetInterface):
    def __init__(self, filename: Path):
        self.filename = Path(filename)
        super().__init__(self.filename.name)

    @contextmanager
    def file(self):
        with open(self.filename, 'rb') as f:
            yield f


@contextmanager
def url_as_file(url: str):
    import urllib.request
    import io
    with urllib.request.urlopen(url) as response:
        xmlx = response.read()
    with io.BytesIO(xmlx) as f:
        yield f


class RemoteSheetInterface(BaseSheetInterface):
    BASE_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    def __init__(self, sheet_id: str, name: str = None):
        super().__init__(name)
        self.sheet_id = sheet_id

    @contextmanager
    def file(self):
        with url_as_file(self.BASE_URL.format(sheet_id=self.sheet_id)) as f:
            yield f


class RemoteListSheetInterface(BaseSheetInterface):
    def __init__(self, sheet_ids: List[str], name: str = None):
        super().__init__(name)
        self.sheet_ids = sheet_ids

    @contextmanager
    def file(self):
        yield None

    def poll(self):
        import openpyxl
        self.tables = []
        for sheet_id in self.sheet_ids:
            with url_as_file(RemoteSheetInterface.BASE_URL.format(sheet_id=sheet_id)) as f:
                wb = openpyxl.load_workbook(f)
                self.tables.append(self._prase_worksheet(wb.active))


class RemoteSheetMultiInterface(MultiInterface):
    def __init__(self, sheet_id: str):
        super().__init__()
        self.sheet_id = sheet_id

    def poll(self):
        import openpyxl
        self.interfaces = []
        with url_as_file(RemoteSheetInterface.BASE_URL.format(sheet_id=self.sheet_id)) as f:
            wb = openpyxl.load_workbook(f)
            for r in wb.active.rows:
                if len(r) > 1:
                    n, *args = r
                    args = [a.value for a in args if a.value is not None]
                    if n.value is not None and len(args) == 1:
                        self.interfaces.append(RemoteSheetInterface(args[0], name=n.value))
                    elif n.value is not None and len(args) > 1:
                        self.interfaces.append(RemoteListSheetInterface(args, name=n.value))
        for i in self.interfaces:
            i.poll()


def main():
    # i = RemoteSheetMultiInterface()
    # i.poll()
    # print(len(i))


if __name__ == "__main__":
    main()
